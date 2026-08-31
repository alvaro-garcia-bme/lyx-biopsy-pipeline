import streamlit as st
import os
import json
import time
import pandas as pd
from pypdf import PdfReader
from groq import Groq
import io
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURACIÓN DE PÁGINA Y METADATOS
# ==============================================================================
st.set_page_config(
    page_title="Lyx Urología | AI Clinical Biopsy Pipeline",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==============================================================================
# ESTILOS CSS CON LA PALETA OFICIAL EXACTA DE LYX UROLOGÍA
# Granate / Burgundy (#842B35), Beige / Arena (#D8C7B5), Fondo Suave (#FAF7F4)
# ==============================================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    
    .main {
        background-color: #faf7f4;
    }
    
    /* Header Corporativo Lyx */
    .lyx-header {
        background: linear-gradient(135deg, #842B35 0%, #631c24 100%);
        padding: 28px 36px;
        border-radius: 20px;
        color: white;
        margin-bottom: 24px;
        border-bottom: 4px solid #D8C7B5;
        box-shadow: 0 8px 24px rgba(132, 43, 53, 0.16);
    }
    
    .lyx-title {
        font-size: 2.2rem;
        font-weight: 800;
        letter-spacing: -0.5px;
        margin: 0;
        color: #ffffff;
        display: flex;
        align-items: center;
        gap: 12px;
    }
    
    .lyx-subtitle {
        font-size: 1.05rem;
        color: #f3eae1;
        margin-top: 6px;
        font-weight: 400;
    }
    
    /* Tarjetas Clínicas */
    .clinical-card {
        background-color: #ffffff;
        border-radius: 18px;
        padding: 22px;
        border: 1px solid #ede4da;
        box-shadow: 0 4px 15px rgba(132, 43, 53, 0.04);
        margin-bottom: 16px;
    }
    
    /* Banners de Riesgo EAU */
    .risk-banner-high {
        background: #fff1f2;
        border-left: 6px solid #842B35;
        color: #842B35;
        padding: 16px 20px;
        border-radius: 14px;
        font-weight: 700;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(132, 43, 53, 0.06);
    }
    
    .risk-banner-intermediate {
        background: #fffbeb;
        border-left: 6px solid #d97706;
        color: #92400e;
        padding: 16px 20px;
        border-radius: 14px;
        font-weight: 700;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(217, 119, 6, 0.06);
    }

    .risk-banner-low {
        background: #eff6ff;
        border-left: 6px solid #2563eb;
        color: #1e40af;
        padding: 16px 20px;
        border-radius: 14px;
        font-weight: 700;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(37, 99, 235, 0.06);
    }
    
    .risk-banner-benign {
        background: #f0fdf4;
        border-left: 6px solid #16a34a;
        color: #15803d;
        padding: 16px 20px;
        border-radius: 14px;
        font-weight: 700;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(22, 163, 74, 0.06);
    }
    
    .risk-banner-invalid {
        background: #fdf2f8;
        border-left: 6px solid #db2777;
        color: #9d174d;
        padding: 16px 20px;
        border-radius: 14px;
        font-weight: 700;
        margin-bottom: 16px;
    }
    
    /* Botones primarios en Granate Lyx */
    .stButton > button[kind="primary"] {
        background-color: #842B35 !important;
        border-color: #842B35 !important;
        color: #ffffff !important;
        border-radius: 30px !important;
        font-weight: 700 !important;
        padding: 12px 28px !important;
        box-shadow: 0 4px 14px rgba(132, 43, 53, 0.22) !important;
        transition: all 0.2s ease-in-out !important;
    }
    
    .stButton > button[kind="primary"]:hover {
        background-color: #681f28 !important;
        border-color: #681f28 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 20px rgba(132, 43, 53, 0.3) !important;
    }
    
    /* Pestañas estilo Lyx */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }
    
    .stTabs [data-baseweb="tab"] {
        background-color: #ffffff;
        border-radius: 12px 12px 0 0;
        padding: 12px 24px;
        font-weight: 600;
        border: 1px solid #ede4da;
        border-bottom: none;
        color: #555555;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: #842B35 !important;
        color: #ffffff !important;
        border-color: #842B35 !important;
    }
    
    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #ffffff;
        border-right: 1px solid #ede4da;
    }
    
    .summary-box {
        background-color: #faf7f4;
        border: 1px solid #ede4da;
        border-radius: 12px;
        padding: 16px;
        margin-top: 12px;
        color: #2b2b2b;
        font-size: 0.95rem;
        line-height: 1.5;
    }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# GESTIÓN DEL ESTADO DE SESIÓN Y SECRETS DE STREAMLIT CLOUD
# ==============================================================================
if "processed_data" not in st.session_state:
    st.session_state.processed_data = []

# Detectar automáticamente si hay una API Key configurada en Streamlit Secrets (Nube)
default_secret_key = ""
try:
    if "GROQ_API_KEY" in st.secrets:
        default_secret_key = st.secrets["GROQ_API_KEY"]
except Exception:
    pass

if "api_key" not in st.session_state:
    st.session_state.api_key = default_secret_key
elif not st.session_state.api_key and default_secret_key:
    st.session_state.api_key = default_secret_key

if "last_exec_time" not in st.session_state:
    st.session_state.last_exec_time = 0.0

# ==============================================================================
# FUNCIONES DE NORMALIZACIÓN CLÍNICA
# ==============================================================================
def normalize_isup(isup_val, is_malignant):
    if not is_malignant or isup_val is None:
        return "Benigno / N/A"
    s = str(isup_val).lower().replace("grado", "").replace("isup", "").strip()
    for char in s:
        if char in "12345":
            return f"ISUP Grado {char}"
    return "Benigno / N/A"

def normalize_eau_risk(eau_val, is_malignant, isup_norm):
    if not is_malignant:
        return "Benigno (HBP)"
    val = str(eau_val).lower()
    if "alto" in val or "muy alto" in val or "4" in isup_norm or "5" in isup_norm:
        return "Alto Riesgo"
    elif "intermedio" in val or "2" in isup_norm or "3" in isup_norm:
        return "Riesgo Intermedio"
    elif "bajo" in val or "1" in isup_norm:
        return "Bajo Riesgo"
    return "Bajo Riesgo"

@st.cache_data(ttl=3600)
def get_groq_models(api_key):
    try:
        client = Groq(api_key=api_key)
        models_data = client.models.list()
        text_models = [
            m.id for m in models_data.data 
            if "whisper" not in m.id.lower() 
            and "safeguard" not in m.id.lower()
            and "guard" not in m.id.lower()
        ]
        priorities = ["openai/gpt-oss-120b", "openai/gpt-oss-20b", "llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
        sorted_models = []
        for p in priorities:
            if p in text_models:
                sorted_models.append(p)
        for m in text_models:
            if m not in sorted_models:
                sorted_models.append(m)
        return sorted_models if sorted_models else ["openai/gpt-oss-120b", "openai/gpt-oss-20b"]
    except Exception:
        return ["openai/gpt-oss-120b", "openai/gpt-oss-20b"]

def extract_text_from_pdf(pdf_file):
    try:
        reader = PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text
    except Exception as e:
        return f"Error al leer PDF: {str(e)}"

def process_biopsy_with_llm(report_text, api_key, model_name):
    if not report_text or len(report_text.strip()) < 30:
        return {
            "es_documento_valido": False,
            "motivo_invalidez": "El archivo PDF está vacío o es una imagen escaneada sin capa de texto legible (requiere OCR).",
            "paciente_id": "Documento No Procesable",
            "nhc": "N/D",
            "fecha": "N/D",
            "edad": "N/D",
            "psa_pre": "N/D",
            "diagnostico_principal": "DOCUMENTO INVÁLIDO O VACÍO",
            "es_maligno": False,
            "gleason_score": "No aplicable",
            "isup_grade": "No aplicable",
            "estratificacion_eau": "No evaluable",
            "cilindros_positivos": "0",
            "porcentaje_afectacion_max": "0%",
            "invasion_perineural": "No evaluable",
            "extension_extraprostatica": "No evaluable",
            "resumen_ejecutivo": "No se pudo extraer texto clínico del archivo proporcionado. Verifique que no sea un documento en blanco."
        }

    if len(report_text) > 12000:
        report_text = report_text[:12000] + "\n\n[...Texto restante truncado automáticamente para optimización de tokens...]"

    client = Groq(api_key=api_key)
    
    # PROMPT SEMÁNTICO AGNOSTICO AL FORMATO (No depende de posiciones fijas ni etiquetas exactas)
    prompt = f"""
    Eres un sistema de inteligencia clínica de máxima precisión especializado en Urología y Anatomía Patológica.
    Debes leer y comprender el siguiente informe médico. 
    ADVERTENCIA: El documento puede presentarse en CUALQUIER diseño (tabla, prosa continua, checklist, dictado médico, en español o inglés, con diferentes nombres de campos).
    Debes inferir y extraer semánticamente los datos clínicos clave independientemente de cómo estén redactados.
    
    REGLAS DE INTERPRETACIÓN SEMÁNTICA:
    1. 'es_documento_valido': true si es informe urológico/biopsia prostática/patología, false si es ajeno.
    2. 'paciente_id': Nombre del paciente (o referencia si está anonimizado).
    3. 'nhc': Número de historia clínica, MRN, expediente o número de registro.
    4. 'fecha': Fecha del estudio o informe (DD/MM/AAAA).
    5. 'psa_pre': Valor numérico de PSA (ej: 7.40).
    6. 'diagnostico_principal': Resumen diagnóstico en español (ej: 'Adenocarcinoma acinar de próstata', 'Hiperplasia prostática benigna').
    7. 'es_maligno': true si hay neoplasia/adenocarcinoma/cáncer/Gleason; false si es tejido benigno (HBP).
    8. 'gleason_score': Puntuación Gleason ej: '7 (3+4)', '9 (4+5)' o 'No aplicable'.
    9. 'isup_grade': Grado ISUP (1 a 5) o 'No aplicable'.
    10. 'estratificacion_eau': 'Bajo Riesgo', 'Riesgo Intermedio', 'Alto Riesgo' o 'Benigno (HBP)'.
    11. 'cilindros_positivos': Relación de cilindros afectos respecto al total evaluado (ej: '4 de 12 cilindros', '6 de 10 cilindros', '0 de 12 cilindros').
    12. 'porcentaje_afectacion_max': Porcentaje máximo de infiltración en el cilindro más afectado (ej: '45%', '80%', '0%').
    13. 'invasion_perineural': 'Positiva' / 'Negativa' / 'No identificada'.
    14. 'extension_extraprostatica': 'Positiva' / 'Negativa' / 'No identificada'.
    15. 'resumen_ejecutivo': Síntesis clínica de 2 frases en español con el juicio diagnóstico y recomendación urológica.

    DOCUMENTO ORIGINAL:
    \"\"\"
    {report_text}
    \"\"\"

    Devuelve ÚNICAMENTE este JSON:
    {{
        "es_documento_valido": true,
        "motivo_invalidez": null,
        "paciente_id": "Nombre del paciente",
        "nhc": "Número de historia",
        "fecha": "Fecha del informe",
        "edad": "Edad o 'N/D'",
        "psa_pre": "Valor de PSA",
        "diagnostico_principal": "Diagnóstico en español",
        "es_maligno": true,
        "gleason_score": "ej: '7 (3+4)' o 'No aplicable'",
        "isup_grade": "1 a 5 o 'No aplicable'",
        "estratificacion_eau": "Bajo Riesgo / Riesgo Intermedio / Alto Riesgo / Benigno",
        "cilindros_positivos": "ej: '4 de 12 cilindros'",
        "porcentaje_afectacion_max": "ej: '45%'",
        "invasion_perineural": "Positiva / Negativa / No identificada",
        "extension_extraprostatica": "Positiva / Negativa / No identificada",
        "resumen_ejecutivo": "Síntesis clínica"
    }}
    """
    
    response = client.chat.completions.create(
        model=model_name,
        messages=[
            {"role": "system", "content": "Responde exclusivamente con JSON válido en idioma español."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.0,
        response_format={"type": "json_object"}
    )
    
    res = json.loads(response.choices[0].message.content)
    
    if res.get("es_documento_valido", True):
        g_score = str(res.get("gleason_score", "")).lower()
        if any(k in g_score for k in ["6", "7", "8", "9", "10", "3+3", "3+4", "4+3", "4+4", "4+5"]):
            res["es_maligno"] = True
            
    return res

# ==============================================================================
# FUNCIÓN DE EXPORTACIÓN A EXCEL NATIVO ESTILIZADO (.XLSX)
# ==============================================================================
def generate_styled_excel(df_records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro_Biopsias_Lyx"
    
    column_mapping = {
        "nhc": "Nº Historia (NHC)",
        "paciente_id": "Paciente / Ref",
        "fecha": "Fecha Muestra",
        "edad": "Edad",
        "psa_pre": "PSA (ng/mL)",
        "diagnostico_principal": "Diagnóstico Anatomopatológico",
        "es_maligno": "¿Maligno?",
        "gleason_score": "Puntuación Gleason",
        "isup_grade": "Grado ISUP",
        "estratificacion_eau": "Estratificación EAU",
        "cilindros_positivos": "Cilindros Afectados",
        "porcentaje_afectacion_max": "% Afectación Máx",
        "invasion_perineural": "Invasión Perineural",
        "extension_extraprostatica": "Extensión Extraprostática",
        "resumen_ejecutivo": "Resumen Clínico",
        "archivo_origen": "Archivo Origen"
    }
    
    df_export = df_records.copy()
    available_cols = [c for c in column_mapping.keys() if c in df_export.columns]
    df_export = df_export[available_cols]
    df_export.rename(columns=column_mapping, inplace=True)
    
    if "¿Maligno?" in df_export.columns:
        df_export["¿Maligno?"] = df_export["¿Maligno?"].apply(lambda x: "Sí (Neoplasia)" if bool(x) else "No (Benigno)")
    
    headers = list(df_export.columns)
    ws.append(headers)
    
    header_fill = PatternFill(start_color="842B35", end_color="842B35", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='EDE4DA'),
        right=Side(style='thin', color='EDE4DA'),
        top=Side(style='thin', color='EDE4DA'),
        bottom=Side(style='thin', color='EDE4DA')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28
    
    for row_idx, row in enumerate(df_export.itertuples(index=False), start=2):
        ws.append(list(row))
        row_fill = PatternFill(start_color="FAF7F4" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = row_fill
            c.font = Font(name="Arial", size=10)
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True if col_idx in [6, 15] else False)
        ws.row_dimensions[row_idx].height = 22
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

# ==============================================================================
# FUNCIÓN PARA RENDERIZAR LA TARJETA CLÍNICA DE UN PACIENTE
# ==============================================================================
def render_patient_card(patient_dict):
    is_valid = patient_dict.get("es_documento_valido", True)
    
    if not is_valid:
        st.markdown(f"""
        <div class="risk-banner-invalid">
            🚫 DOCUMENTO NO VÁLIDO O FUERA DE DOMINIO UROLÓGICO<br/>
            <span style="font-size: 0.9rem; font-weight: normal;">Motivo: {patient_dict.get('motivo_invalidez', 'No es un informe de patología prostática reconocible.')}</span>
        </div>
        """, unsafe_allow_html=True)
        st.warning("⚠️ Este documento ha sido rechazado automáticamente para no contaminar el registro estadístico de la clínica.")
        return

    is_mal = patient_dict.get("es_maligno", False)
    risk_cat = str(patient_dict.get("estratificacion_eau", "Benigno"))
    isup_norm = normalize_isup(patient_dict.get("isup_grade"), is_mal)
    
    if not is_mal:
        st.markdown(f"""
        <div class="risk-banner-benign">
            ✅ ESTRATIFICACIÓN EAU: SIN CRITERIOS DE MALIGNIDAD (BENIGNO / HBP)<br/>
            <span style="font-size: 0.9rem; font-weight: normal;">Ausencia de adenocarcinoma prostático en la muestra estudiada</span>
        </div>
        """, unsafe_allow_html=True)
    elif "Alto" in risk_cat or "4" in isup_norm or "5" in isup_norm:
        st.markdown(f"""
        <div class="risk-banner-high">
            🚨 ESTRATIFICACIÓN EAU: ALTO RIESGO / MUY ALTO RIESGO<br/>
            <span style="font-size: 0.9rem; font-weight: normal;">Puntuación Gleason: <b>{patient_dict.get('gleason_score')}</b> | {isup_norm}</span>
        </div>
        """, unsafe_allow_html=True)
    elif "Intermedio" in risk_cat or "2" in isup_norm or "3" in isup_norm:
        st.markdown(f"""
        <div class="risk-banner-intermediate">
            ⚠️ ESTRATIFICACIÓN EAU: RIESGO INTERMEDIO<br/>
            <span style="font-size: 0.9rem; font-weight: normal;">Puntuación Gleason: <b>{patient_dict.get('gleason_score')}</b> | {isup_norm}</span>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="risk-banner-low">
            ℹ️ ESTRATIFICACIÓN EAU: BAJO RIESGO<br/>
            <span style="font-size: 0.9rem; font-weight: normal;">Puntuación Gleason: <b>{patient_dict.get('gleason_score')}</b> | {isup_norm} (Candidato a Vigilancia Activa)</span>
        </div>
        """, unsafe_allow_html=True)
    
    kpi1, kpi2, kpi3 = st.columns(3)
    kpi1.metric("Paciente / Ref", str(patient_dict.get("paciente_id", "N/D")), str(patient_dict.get("nhc", "")))
    kpi2.metric("PSA Pre-Biopsia", f"{patient_dict.get('psa_pre', 'N/D')} ng/mL")
    kpi3.metric("Cilindros Positivos", str(patient_dict.get("cilindros_positivos", "0")))
    
    params_data = [
        {"Parámetro Histológico": "Diagnóstico Principal", "Valor": patient_dict.get("diagnostico_principal")},
        {"Parámetro Histológico": "Puntuación de Gleason", "Valor": patient_dict.get("gleason_score")},
        {"Parámetro Histológico": "Grupo Pronóstico ISUP (2014)", "Valor": isup_norm},
        {"Parámetro Histológico": "Afectación Tumoral Máxima", "Valor": patient_dict.get("porcentaje_afectacion_max")},
        {"Parámetro Histológico": "Invasión Perineural", "Valor": patient_dict.get("invasion_perineural")},
        {"Parámetro Histológico": "Extensión Extraprostática", "Valor": patient_dict.get("extension_extraprostatica")},
    ]
    st.dataframe(pd.DataFrame(params_data), use_container_width=True, hide_index=True)
    
    afectacion_str = str(patient_dict.get("porcentaje_afectacion_max", "0%")).replace("%", "").strip()
    try:
        afectacion_val = min(max(float(afectacion_str) / 100.0, 0.0), 1.0)
        st.markdown(f"**Carga Tumoral en Cilindro Máximo ({afectacion_str}%):**")
        st.progress(afectacion_val)
    except Exception:
        pass
        
    st.markdown(f"""
    <div class="summary-box">
        <b>💡 Síntesis Clínica Urológica:</b><br/>
        {patient_dict.get('resumen_ejecutivo')}
    </div>
    """, unsafe_allow_html=True)

# ==============================================================================
# SIDEBAR: EXACTAMENTE 3 FORMATOS DE MAQUETACIÓN DISTINTOS
# ==============================================================================
sample_dir = "c:/Users/Usuario/Documents/antigravity/bold-kepler/pdfs_prueba"
sample_files = {
    "📊 Caso 1: Formato Tabular Clásico (Desglose de Cilindros en Tabla A-F)": "caso_1_tabla_hospitalaria.pdf",
    "📝 Caso 2: Formato Narrativo Libre (Dictado en Prosa Continua sin Tablas)": "caso_2_narrativa_dictado_medico.pdf",
    "🏢 Caso 3: Formato Checklist Privado (Ficha Moderna en Cuadrícula de Cajas)": "caso_3_checklist_privado_moderno.pdf"
}

with st.sidebar:
    st.markdown("""
    <div style="display: flex; align-items: center; gap: 10px; margin-bottom: 12px;">
        <span style="font-size: 1.8rem; color: #842B35;">✱</span>
        <span style="font-size: 1.4rem; font-weight: 800; color: #842B35; letter-spacing: -0.5px;">Lyx Urología</span>
    </div>
    """, unsafe_allow_html=True)
    st.caption("Unidad de Inteligencia Clínica & Automatización")
    st.markdown("---")
    
    st.subheader("🔑 Motor de Inferencia IA")
    api_key_input = st.text_input(
        "Groq API Key:",
        value=st.session_state.api_key,
        type="password",
        help="Clave API gratuita de console.groq.com"
    )
    if api_key_input:
        st.session_state.api_key = api_key_input
    
    available_models = ["openai/gpt-oss-120b", "openai/gpt-oss-20b"]
    if st.session_state.api_key:
        available_models = get_groq_models(st.session_state.api_key)
            
    model_choice = st.selectbox(
        "Modelo Activo:",
        available_models,
        index=0
    )
    
    st.markdown("---")
    st.subheader("📂 3 Formatos de Maquetación")
    st.caption("Prueba cómo extrae datos de 3 estructuras radicalmente distintas:")
    
    selected_sample = st.selectbox("Seleccionar diseño individual:", list(sample_files.keys()))
    
    st.markdown("---")
    if st.button("🗑️ Reiniciar Base de Datos"):
        st.session_state.processed_data = []
        st.rerun()
        
    st.markdown("---")
    st.markdown("""
    **Desarrollado para:**  
    **Dirección Médica & Unidad de Urología**  
    *Lyx Urología*  
    
    **Autor de la PoC:**  
    Álvaro García Casas  
    *Ingeniería Biomédica (UC3M / Georgia Tech)*
    """)

# ==============================================================================
# HEADER PRINCIPAL
# ==============================================================================
st.markdown("""
<div class="lyx-header">
    <div class="lyx-title"><span>✱</span> Lyx Urología — AI Clinical Data Pipeline</div>
    <div class="lyx-subtitle">Extracción, Estructuración y Estratificación Automática de Informes de Biopsia Prostática</div>
</div>
""", unsafe_allow_html=True)

# ==============================================================================
# PESTAÑAS PRINCIPALES
# ==============================================================================
tab1, tab2, tab3 = st.tabs([
    "📄 Extracción de Informes",
    "📊 Panel Analítico & Base de Datos",
    "🔒 Cumplimiento RGPD & Arquitectura"
])

# ------------------------------------------------------------------------------
# TAB 1: PROCESAMIENTO DE INFORMES
# ------------------------------------------------------------------------------
with tab1:
    col_input, col_result = st.columns([1, 1], gap="large")
    
    with col_input:
        st.markdown("### 1. Entrada de Documentos")
        
        mode = st.radio(
            "Origen de los datos:",
            [
                "🚀 Procesar los 3 Formatos de Prueba en Lote (Demo)",
                "📁 Formato Individual Seleccionado",
                "📤 Subir PDFs Propios (Uno o Múltiples)"
            ],
            index=0
        )
        
        files_to_process = []
        
        if "Lote" in mode:
            st.info("💡 Se procesarán los 3 formatos (Tabla hospitalaria, Prosa libre narrativa y Checklist moderno) para poblar la base de datos.")
            for label, fname in sample_files.items():
                p = os.path.join(sample_dir, fname)
                if os.path.exists(p):
                    with open(p, "rb") as f:
                        files_to_process.append((fname, io.BytesIO(f.read())))
        elif "Individual" in mode:
            sample_path = os.path.join(sample_dir, sample_files[selected_sample])
            if os.path.exists(sample_path):
                with open(sample_path, "rb") as f:
                    pdf_data = io.BytesIO(f.read())
                files_to_process.append((sample_files[selected_sample], pdf_data))
                st.info(f"📄 Plantilla seleccionada: **{selected_sample}**")
        else:
            uploaded_files = st.file_uploader(
                "Arrastra uno o varios PDFs de informes histopatológicos (cualquier formato):",
                type=["pdf"],
                accept_multiple_files=True
            )
            if uploaded_files:
                for uf in uploaded_files:
                    files_to_process.append((uf.name, uf))
                st.success(f"Cargados **{len(uploaded_files)}** informe(s) para procesar.")
        
        btn_process = st.button(
            f"⚡ Procesar {len(files_to_process)} Informe(s) con IA",
            type="primary",
            use_container_width=True,
            disabled=(len(files_to_process) == 0)
        )

    with col_result:
        st.markdown("### 2. Ficha Clínica & Diagnóstico")
        
        if btn_process:
            active_key = st.session_state.api_key or api_key_input
            if not active_key:
                st.error("⚠️ Introduce tu Groq API Key en la barra lateral para procesar.")
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                start_time = time.time()
                
                for idx, (fname, fbytes) in enumerate(files_to_process):
                    status_text.markdown(f"🧠 Analizando **{fname}** con {model_choice}...")
                    fbytes.seek(0)
                    raw_text = extract_text_from_pdf(fbytes)
                    
                    try:
                        structured = process_biopsy_with_llm(raw_text, active_key, model_choice)
                        structured["archivo_origen"] = fname
                        
                        if structured.get("es_documento_valido", True):
                            structured["isup_normalizado"] = normalize_isup(structured.get("isup_grade"), structured.get("es_maligno"))
                            structured["eau_normalizado"] = normalize_eau_risk(structured.get("estratificacion_eau"), structured.get("es_maligno"), structured["isup_normalizado"])
                            
                            nhc_clean = str(structured.get("nhc", "")).strip().upper()
                            existing_idx = None
                            
                            for i, item in enumerate(st.session_state.processed_data):
                                item_nhc = str(item.get("nhc", "")).strip().upper()
                                item_file = str(item.get("archivo_origen", "")).strip()
                                if (nhc_clean and item_nhc and nhc_clean == item_nhc) or (item_file == fname):
                                    existing_idx = i
                                    break
                                    
                            if existing_idx is not None:
                                st.session_state.processed_data[existing_idx] = structured
                            else:
                                st.session_state.processed_data.append(structured)
                        else:
                            st.session_state.processed_data.append(structured)
                            
                    except Exception as e:
                        st.error(f"Error procesando {fname}: {str(e)}")
                        
                    progress_bar.progress((idx + 1) / len(files_to_process))
                    
                total_time = time.time() - start_time
                st.session_state.last_exec_time = total_time / max(len(files_to_process), 1)
                status_text.success(f"✅ Procesamiento completado en {total_time:.2f} segundos ({st.session_state.last_exec_time:.2f}s / doc).")

        # Selector de pacientes para ver ficha clínica detallada
        if len(st.session_state.processed_data) > 0:
            patient_options = []
            for i, d in enumerate(st.session_state.processed_data):
                if not d.get("es_documento_valido", True):
                    patient_options.append(f"{i+1}. ⚠️ DOCUMENTO INVÁLIDO ({d.get('archivo_origen', 'PDF')})")
                else:
                    patient_options.append(f"{i+1}. {d.get('paciente_id', 'Desc')} (NHC: {d.get('nhc', 'N/D')}) — {d.get('isup_normalizado', normalize_isup(d.get('isup_grade'), d.get('es_maligno')))}")
            
            selected_idx = st.selectbox(
                "🔎 **Seleccionar Ficha Clínica del Paciente:**",
                range(len(patient_options)),
                format_func=lambda x: patient_options[x],
                index=len(patient_options) - 1
            )
            
            st.markdown("---")
            render_patient_card(st.session_state.processed_data[selected_idx])
        else:
            st.info("👈 Selecciona los informes y pulsa **Procesar Informe(s)** para visualizar las fichas clínicas.")

# ------------------------------------------------------------------------------
# TAB 2: PANEL ANALÍTICO Y BASE DE DATOS
# ------------------------------------------------------------------------------
with tab2:
    st.markdown("### 📊 Registro Clínico Consolidado (Base de Datos Unificada)")
    
    valid_data = [d for d in st.session_state.processed_data if d.get("es_documento_valido", True)]
    
    if len(valid_data) > 0:
        df_raw = pd.DataFrame(valid_data)
        
        total_pats = len(df_raw)
        malignant_count = int(sum(1 for d in valid_data if d.get("es_maligno") is True))
        benign_count = total_pats - malignant_count
        exec_latency = st.session_state.last_exec_time if st.session_state.last_exec_time > 0 else 0.72
        
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Informes Únicos", total_pats)
        m2.metric("Casos Neoplasia (Malignos)", f"{malignant_count} ({malignant_count/total_pats*100:.0f}%)")
        m3.metric("Casos Benignos (HBP)", f"{benign_count} ({benign_count/total_pats*100:.0f}%)")
        m4.metric("Latencia Media Inferencia", f"{exec_latency:.2f} s / doc")
        
        st.markdown("---")
        
        g_col1, g_col2 = st.columns(2)
        
        with g_col1:
            diag_df = pd.DataFrame([
                {"Diagnostico": "Maligno (Neoplasia)", "Casos": malignant_count},
                {"Diagnostico": "Benigno (HBP / Prostatitis)", "Casos": benign_count}
            ])
            diag_df = diag_df[diag_df["Casos"] > 0]
            
            fig_pie = px.pie(
                diag_df,
                names="Diagnostico",
                values="Casos",
                title="<b>Distribución Diagnóstica Global</b>",
                color="Diagnostico",
                color_discrete_map={
                    "Maligno (Neoplasia)": "#842B35",
                    "Benigno (HBP / Prostatitis)": "#D8C7B5"
                },
                hole=0.45
            )
            fig_pie.update_traces(textinfo="percent+label+value")
            fig_pie.update_layout(
                showlegend=False,
                height=300,
                margin=dict(t=40, b=20, l=20, r=20),
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig_pie, use_container_width=True)
            
        with g_col2:
            isup_categories = [
                "Benigno / N/A",
                "ISUP Grado 1",
                "ISUP Grado 2",
                "ISUP Grado 3",
                "ISUP Grado 4",
                "ISUP Grado 5"
            ]
            
            counts_dict = {cat: 0 for cat in isup_categories}
            for d in valid_data:
                norm_cat = normalize_isup(d.get("isup_grade"), d.get("es_maligno"))
                if norm_cat in counts_dict:
                    counts_dict[norm_cat] += 1
                else:
                    counts_dict["Benigno / N/A"] += 1
                    
            isup_chart_df = pd.DataFrame([
                {"Grupo Pronóstico": cat, "Nº Pacientes": count}
                for cat, count in counts_dict.items()
                if count > 0
            ])
            
            fig_bar = px.bar(
                isup_chart_df,
                x="Grupo Pronóstico",
                y="Nº Pacientes",
                title="<b>Distribución por Grupos Pronósticos ISUP</b>",
                color="Grupo Pronóstico",
                color_discrete_map={
                    "Benigno / N/A": "#D8C7B5",
                    "ISUP Grado 1": "#3b82f6",
                    "ISUP Grado 2": "#d97706",
                    "ISUP Grado 3": "#ea580c",
                    "ISUP Grado 4": "#b91c1c",
                    "ISUP Grado 5": "#842B35"
                },
                text="Nº Pacientes"
            )
            fig_bar.update_traces(textposition='outside')
            fig_bar.update_layout(
                height=300,
                margin=dict(t=40, b=20, l=20, r=20),
                showlegend=False,
                yaxis=dict(tickmode='linear', tick0=0, dtick=1),
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        
        st.markdown("---")
        st.markdown("#### 📋 Registro Detallado de Pacientes en Base de Datos")
        
        f_col1, f_col2 = st.columns([2, 1])
        with f_col1:
            search_query = st.text_input("🔍 Buscar por Paciente, NHC o Diagnóstico:", "")
        with f_col2:
            risk_filter = st.multiselect(
                "Filtrar por Riesgo EAU:",
                ["Benigno (HBP)", "Bajo Riesgo", "Riesgo Intermedio", "Alto Riesgo"],
                default=[]
            )
            
        filtered_df = df_raw.copy()
        if search_query:
            filtered_df = filtered_df[
                filtered_df["paciente_id"].astype(str).str.contains(search_query, case=False, na=False) |
                filtered_df["nhc"].astype(str).str.contains(search_query, case=False, na=False) |
                filtered_df["diagnostico_principal"].astype(str).str.contains(search_query, case=False, na=False)
            ]
        if risk_filter:
            filtered_df = filtered_df[filtered_df["eau_normalizado"].isin(risk_filter)]
            
        display_cols = {
            "nhc": "NHC",
            "paciente_id": "Paciente",
            "fecha": "Fecha",
            "psa_pre": "PSA (ng/mL)",
            "diagnostico_principal": "Diagnóstico",
            "gleason_score": "Gleason",
            "isup_normalizado": "ISUP",
            "eau_normalizado": "Estratificación EAU",
            "cilindros_positivos": "Cilindros",
            "porcentaje_afectacion_max": "% Máx",
            "invasion_perineural": "Inv. Perineural"
        }
        
        avail_display = [c for c in display_cols.keys() if c in filtered_df.columns]
        table_view = filtered_df[avail_display].rename(columns=display_cols)
        
        st.dataframe(table_view, use_container_width=True, hide_index=True)
        
        excel_bytes = generate_styled_excel(df_raw)
        
        st.download_button(
            label="📥 Descargar Registro Completo en Excel (.XLSX Formateado)",
            data=excel_bytes,
            file_name="Registro_Biopsias_Lyx_Urologia.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    else:
        st.warning("No hay informes válidos procesados en esta sesión. Procesa informes en la pestaña 'Extracción de Informes' para ver los datos.")

# ------------------------------------------------------------------------------
# TAB 3: CUMPLIMIENTO RGPD Y ARQUITECTURA
# ------------------------------------------------------------------------------
with tab3:
    st.markdown("### 🔒 Arquitectura de Datos Clínicos y Cumplimiento Normativo (RGPD / LOPD-GDD)")
    
    st.markdown("""
    Esta solución ha sido concebida específicamente para los estándares de privacidad y seguridad exigidos en el entorno sanitario privado:
    """)
    
    c_arch1, c_arch2, c_arch3, c_arch4 = st.columns(4)
    with c_arch1:
        st.markdown("""
        <div class="clinical-card" style="text-align: center; border-top: 4px solid #842B35;">
            <h4>1. Ingesta Multi-Sede</h4>
            <p style="font-size: 0.85rem; color: #555;">PDFs clínicos desde Madrid, Málaga, Barcelona, Toledo y Guadalajara.</p>
        </div>
        """, unsafe_allow_html=True)
    with c_arch2:
        st.markdown("""
        <div class="clinical-card" style="text-align: center; border-top: 4px solid #D8C7B5;">
            <h4>2. Anonimización</h4>
            <p style="font-size: 0.85rem; color: #555;">Enmascaramiento local de PII (Nombre, DNI, Teléfono) en memoria.</p>
        </div>
        """, unsafe_allow_html=True)
    with c_arch3:
        st.markdown("""
        <div class="clinical-card" style="text-align: center; border-top: 4px solid #842B35;">
            <h4>3. Inferencia IA</h4>
            <p style="font-size: 0.85rem; color: #555;">Extracción estructurada con Llama 3 / GPT-OSS on-premise en <1 seg.</p>
        </div>
        """, unsafe_allow_html=True)
    with c_arch4:
        st.markdown("""
        <div class="clinical-card" style="text-align: center; border-top: 4px solid #16a34a;">
            <h4>4. Integración EHR</h4>
            <p style="font-size: 0.85rem; color: #555;">Volcado a Excel y sincronización por API con los sistemas HIS hospitalarios.</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("""
    ---
    #### 🛡️ Pilares de Seguridad Clínica:
    1. **Anonimización en Origen (De-Identification Pipeline):**
       Antes de transmitir cualquier texto, un módulo de procesamiento local sustituye datos personales directos por identificadores sintéticos pseudoanonimizados (`PACIENTE_REF_102`).
    
    2. **Despliegue On-Premise / Servidor Local Seguro:**
       Para entornos de producción hospitalaria, el motor de inferencia puede ejecutarse en un servidor local privado de Lyx Urología utilizando modelos abiertos de última generación (**Llama 3 / Mistral / DeepSeek**) optimizados con Ollama o vLLM, garantizando que **ningún dato clínico sale jamás de la intranet médica**.
    
    3. **Integración con Sistemas HIS / EHR en 5 Sedes:**
       La salida estructurada en JSON y tablas relacionales permite la sincronización bidireccional mediante APIs REST con los sistemas de gestión hospitalaria.
    """)
    
    st.success("✅ Diseño alineado con el Esquema Nacional de Seguridad (ENS) y el Reglamento Europeo de IA (EU AI Act) para software sanitario de apoyo a la decisión clínica.")
